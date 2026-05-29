#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Data Listing Parser for AE Risk Assessment Skill.
Auto-detects sheet types and extracts lab, vital signs, ECG, AE, MH, CM data.
"""
import json
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl


def norm(s):
    """Normalize string: strip, convert to str."""
    if s is None:
        return ""
    if isinstance(s, (int, float)):
        return str(s)
    return str(s).strip()


def to_float(x):
    """Try to convert to float; return None on failure."""
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    try:
        return float(norm(x))
    except (ValueError, TypeError):
        return None


def to_date(x):
    """Try to convert to date string."""
    if x is None:
        return None
    if hasattr(x, "strftime"):
        return x.strftime("%Y-%m-%d")
    s = norm(x)
    if not s or s.lower() == "nan":
        return None
    patterns = [
        (r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", "%Y-%m-%d"),
        (r"(\d{1,2})/(\d{1,2})/(\d{4})", "%m/%d/%Y"),
        (r"(\d{4})年(\d{1,2})月(\d{1,2})日", "%Y-%m-%d"),
    ]
    for pat, fmt in patterns:
        m = re.search(pat, s)
        if m:
            if "%Y-%m-%d" == fmt:
                return f"{m.group(1)}-{m.group(2).zfill(2)}-{m.group(3).zfill(2)}"
            else:
                groups = m.groups()
                try:
                    dt = datetime.strptime(f"{groups[0]}/{groups[1]}/{groups[2]}", fmt)
                    return dt.strftime("%Y-%m-%d")
                except ValueError:
                    continue
    try:
        return datetime.strptime(s[:10], "%Y-%m-%d").strftime("%Y-%m-%d")
    except ValueError:
        pass
    return s


def extract_visit_point(visit_name):
    """Extract V{n} from visit name like '筛选/导入期V1（D-7~D-1）' → 'V1'.
    Only returns V{n} if the visit name actually contains it. Returns '' otherwise."""
    s = norm(visit_name)
    m = re.search(r'V(\d+)', s)
    if m:
        return f"V{m.group(1)}"
    return ""


def col_index(headers, *candidates):
    """Find the first matching column name from candidates."""
    for h, idx in headers.items():
        h_norm = norm(h)
        for c in candidates:
            if c.lower() in h_norm.lower():
                return idx
    return None


def match_col(headers, *patterns):
    """Match column name by regex patterns."""
    for h, idx in headers.items():
        h_norm = norm(h)
        for p in patterns:
            if re.search(p, h_norm, re.IGNORECASE):
                return idx
    return None


def classify_sheet(name, headers_list):
    """Classify a sheet based on its name and headers."""
    name_lower = name.lower()
    header_text = " ".join(str(h or "") for h in headers_list).lower()

    # Visit sheet
    if any(k in name_lower for k in ["sv", "访视"]) and "sv" not in name_lower.replace("sv", ""):
        return "SV"

    if "dm" == name_lower.strip() or "人口" in name_lower:
        return "DM"

    # Medical History
    if any(k in name_lower for k in ["mh", "病史", "既往及现病史"]):
        return "MH"

    # Adverse Events
    if any(k in name_lower for k in ["ae", "不良事件"]):
        return "AE"

    # Concomitant Medications
    if any(k in name_lower for k in ["cm", "合并用药"]):
        return "CM"

    # Lab - Hematology
    if "hem" in name_lower or "血常规" in name_lower:
        return "LB_HEM"

    # Lab - Chemistry
    if "chem" in name_lower or "血生化" in name_lower:
        return "LB_CHEM"

    # Lab - Urinalysis
    if "uri" in name_lower or "尿常规" in name_lower:
        return "LB_URI"

    # Lab - Other (HBV, HCV, HCG, VIR, etc.)
    if name_lower.startswith("lb_"):
        return "LB_OTHER"

    # Vital Signs
    if any(k in name_lower for k in ["vs", "生命体征", "vsw"]):
        return "VS"

    # ECG
    if any(k in name_lower for k in ["eg", "心电图", "ecg"]):
        return "EG"

    # Height/Weight
    if any(k in name_lower for k in ["hw", "身高体重", "体格检查"]):
        return "HW"

    # Subject groupings
    if "分组" in name_lower or "治疗组" in name_lower:
        return "GROUP"

    return None


def parse_subject_id(val):
    """Extract clean subject ID."""
    s = norm(val)
    if not s or s.lower() == "nan":
        return None
    # Remove common prefixes
    s = re.sub(r'^[A-Za-z]+[-_ ]', '', s)
    return s.strip()


def parse_lab_rows(ws, sheet_type):
    """Parse lab test rows from a sheet."""
    rows_data = list(ws.iter_rows(values_only=True))
    if not rows_data:
        return []
    headers_raw = [norm(h) for h in rows_data[0]]
    headers = {h: i for i, h in enumerate(headers_raw) if h}

    subj_col = col_index(headers, "受试者编号", "受试者", "SUBJID", "Subject")
    center_col = col_index(headers, "试验中心名称", "中心名称", "研究中心", "Site")
    center_id_col = col_index(headers, "试验中心编号", "中心编号")
    date_col = col_index(headers, "采样日期", "检查日期", "LBDAT")
    # Priority: 数据节 > 访视 > VISIT (per SKILL.md 4.3b)
    visit_col = col_index(headers, "数据节")
    if visit_col is None:
        visit_col = match_col(headers, r"访视|VISIT")
    visit_point_col = match_col(headers, r"访视点|访视窗|时间点|TPT|VISITPOINT|INSTANCE")
    test_col = col_index(headers, "实验室指标名称", "检查项目", "指标名称", "LBTEST")
    result_col = col_index(headers, "结果", "检查结果", "LBORRES", "VSORRES")
    unit_col = col_index(headers, "单位", "LBORRESU", "VSORRESU")
    lo_col = col_index(headers, "下限", "正常值范围下限", "LBORNRL")
    hi_col = col_index(headers, "上限", "正常值范围上限", "LBORNRH")
    cs_col = col_index(headers, "临床评估", "临床意义", "LBCLSIG", "VSCLSIG")
    desc_col = col_index(headers, "异常说明", "若异常有临床意义，请说明",
                         "若异常，请详述", "异常，请说明")
    desc_link_col = col_index(headers, "动态链接")
    note_col = col_index(headers, "备注", "LBCO", "VSCO", "EGCO")
    ae_col = col_index(headers, "不良事件", "AE", "AENO", "VSAENO")
    mh_col = col_index(headers, "其他既往及现病史", "MHNO")
    perf_col = col_index(headers, "是否检查", "是否评估", "是否测量")

    records = []
    for row in rows_data[1:]:
        subj = parse_subject_id(row[subj_col]) if subj_col is not None else None
        if not subj:
            continue

        # Check if test was performed
        if perf_col is not None:
            perf_val = norm(row[perf_col])
            if perf_val in ("否", "No", "no", "N", "n", "0"):
                continue

        test_name = norm(row[test_col]) if test_col is not None else ""
        if not test_name:
            continue

        result = to_float(row[result_col]) if result_col is not None else None
        lo = to_float(row[lo_col]) if lo_col is not None else None
        hi = to_float(row[hi_col]) if hi_col is not None else None
        unit = norm(row[unit_col]) if unit_col is not None else ""
        cs = norm(row[cs_col]) if cs_col is not None else ""
        desc = norm(row[desc_col]) if desc_col is not None else ""
        desc_link = norm(row[desc_link_col]) if desc_link_col is not None else ""
        note = norm(row[note_col]) if note_col is not None else ""
        date = to_date(row[date_col]) if date_col is not None else None
        visit = norm(row[visit_col]) if visit_col is not None else date or ""
        visit_point = extract_visit_point(visit)
        center = norm(row[center_col]) if center_col is not None else (
            norm(row[center_id_col]) if center_id_col is not None else ""
        )
        ae_ref = norm(row[ae_col]) if ae_col is not None else ""
        mh_ref = norm(row[mh_col]) if mh_col is not None else ""

        record = {
            "subj": subj,
            "center": center,
            "visit": visit,
            "visit_point": visit_point,
            "date": date,
            "test": test_name,
            "result": result,
            "unit": unit,
            "lo": lo,
            "hi": hi,
            "cs": cs,
            "desc": desc,
            "desc_link": desc_link,
            "note": note,
            "ae_ref": ae_ref,
            "mh_ref": mh_ref,
            "source_sheet": sheet_type,
        }
        records.append(record)

    return records


def parse_vs_rows(ws):
    """Parse vital signs rows (VS sheet)."""
    rows_data = list(ws.iter_rows(values_only=True))
    if not rows_data:
        return []
    headers_raw = [norm(h) for h in rows_data[0]]
    headers = {h: i for i, h in enumerate(headers_raw) if h}

    subj_col = col_index(headers, "受试者编号", "受试者", "SUBJID")
    center_col = col_index(headers, "试验中心名称", "中心名称", "研究中心")
    center_id_col = col_index(headers, "试验中心编号", "中心编号")
    date_col = col_index(headers, "检查日期", "测量日期", "VSDAT")
    visit_col = col_index(headers, "数据节")
    if visit_col is None:
        visit_col = match_col(headers, r"访视|VISIT")
    visit_point_col = match_col(headers, r"访视点|访视窗|时间点|TPT|VISITPOINT|INSTANCE")
    test_col = col_index(headers, "检查项目", "VSTEST")
    result_col = col_index(headers, "检查结果", "VSORRES", "结果")
    unit_col = col_index(headers, "单位", "VSORRESU")
    cs_col = col_index(headers, "临床评估", "VSCLSIG")
    desc_col = col_index(headers, "异常说明", "异常，请说明", "若异常有临床意义，请说明")
    ae_col = col_index(headers, "不良事件", "VSAENO")
    mh_col = col_index(headers, "其他既往及现病史", "VSMHNO")
    perf_col = col_index(headers, "是否检查")
    note_col = col_index(headers, "备注", "VSCO")

    records = []
    for row in rows_data[1:]:
        subj = parse_subject_id(row[subj_col]) if subj_col is not None else None
        if not subj:
            continue
        if perf_col is not None and norm(row[perf_col]) in ("否", "No", "no"):
            continue
        test_name = norm(row[test_col]) if test_col is not None else ""
        if not test_name:
            continue
        result = to_float(row[result_col]) if result_col is not None else None
        unit = norm(row[unit_col]) if unit_col is not None else ""
        cs = norm(row[cs_col]) if cs_col is not None else ""
        desc = norm(row[desc_col]) if desc_col is not None else ""
        date = to_date(row[date_col]) if date_col is not None else None
        visit = norm(row[visit_col]) if visit_col is not None else date or ""
        visit_point = extract_visit_point(visit)
        center = norm(row[center_col]) if center_col is not None else (
            norm(row[center_id_col]) if center_id_col is not None else ""
        )
        records.append({
            "subj": subj, "center": center, "visit": visit, "visit_point": visit_point, "date": date,
            "test": test_name, "result": result, "unit": unit,
            "lo": None, "hi": None, "cs": cs, "desc": desc,
            "desc_link": "", "note": norm(row[note_col]) if note_col is not None else "",
            "ae_ref": norm(row[ae_col]) if ae_col is not None else "",
            "mh_ref": norm(row[mh_col]) if mh_col is not None else "",
            "source_sheet": "VS",
        })
    return records


def parse_eg_rows(ws):
    """Parse ECG rows."""
    rows_data = list(ws.iter_rows(values_only=True))
    if not rows_data:
        return []
    headers_raw = [norm(h) for h in rows_data[0]]
    headers = {h: i for i, h in enumerate(headers_raw) if h}

    subj_col = col_index(headers, "受试者编号", "受试者", "SUBJID")
    center_col = col_index(headers, "试验中心名称", "中心名称", "研究中心")
    center_id_col = col_index(headers, "试验中心编号", "中心编号")
    date_col = col_index(headers, "检查日期", "EGDAT")
    visit_col = col_index(headers, "数据节")
    if visit_col is None:
        visit_col = match_col(headers, r"访视|VISIT")
    visit_point_col = match_col(headers, r"访视点|访视窗|时间点|TPT|VISITPOINT|INSTANCE")
    cs_col = col_index(headers, "临床评估", "EGCLSIG")
    desc_col = col_index(headers, "异常说明", "若异常有临床意义，请说明",
                         "若异常，请详述")
    note_col = col_index(headers, "备注", "EGCO")
    ae_col = col_index(headers, "不良事件", "EGAENO")
    mh_col = col_index(headers, "其他既往及现病史", "EGMHNO")
    perf_col = col_index(headers, "是否检查")

    # ECG test columns
    ecg_tests = {
        "QTcF": col_index(headers, "QTcF", "QTCF"),
        "QTcB": col_index(headers, "QTcB", "QTCB"),
        "QT间期": col_index(headers, "QT间期", "EGQT"),
        "心率": col_index(headers, "心率", "EGHR"),
        "PR间期": col_index(headers, "PR间期", "EGPR"),
        "QRS": col_index(headers, "QRS", "EGQRS"),
        "RR间期": col_index(headers, "RR间期", "EGRR"),
    }

    records = []
    for row in rows_data[1:]:
        subj = parse_subject_id(row[subj_col]) if subj_col is not None else None
        if not subj:
            continue
        if perf_col is not None and norm(row[perf_col]) in ("否", "No", "no"):
            continue
        date = to_date(row[date_col]) if date_col is not None else None
        visit = norm(row[visit_col]) if visit_col is not None else date or ""
        visit_point = extract_visit_point(visit)
        center = norm(row[center_col]) if center_col is not None else (
            norm(row[center_id_col]) if center_id_col is not None else ""
        )
        cs = norm(row[cs_col]) if cs_col is not None else ""
        desc = norm(row[desc_col]) if desc_col is not None else ""
        note = norm(row[note_col]) if note_col is not None else ""
        ae_ref = norm(row[ae_col]) if ae_col is not None else ""
        mh_ref = norm(row[mh_col]) if mh_col is not None else ""

        for test_name, test_col in ecg_tests.items():
            if test_col is None:
                continue
            result = to_float(row[test_col])
            if result is not None:
                records.append({
                    "subj": subj, "center": center, "visit": visit, "visit_point": visit_point, "date": date,
                    "test": test_name, "result": result, "unit": "ms" if "QT" in test_name or "QRS" in test_name or "PR" in test_name or "RR" in test_name else "bpm",
                    "lo": None, "hi": None, "cs": cs, "desc": desc,
                    "desc_link": "", "note": note,
                    "ae_ref": ae_ref, "mh_ref": mh_ref,
                    "source_sheet": "EG",
                })
    return records


def parse_hw_rows(ws):
    """Parse height/weight rows."""
    rows_data = list(ws.iter_rows(values_only=True))
    if not rows_data:
        return []
    headers_raw = [norm(h) for h in rows_data[0]]
    headers = {h: i for i, h in enumerate(headers_raw) if h}

    subj_col = col_index(headers, "受试者编号", "受试者", "SUBJID")
    center_col = col_index(headers, "试验中心名称", "中心名称", "研究中心")
    center_id_col = col_index(headers, "试验中心编号", "中心编号")
    date_col = col_index(headers, "测量日期", "检查日期", "HWDAT")
    visit_col = col_index(headers, "数据节")
    if visit_col is None:
        visit_col = match_col(headers, r"访视|VISIT")
    visit_point_col = match_col(headers, r"访视点|访视窗|时间点|TPT|VISITPOINT|INSTANCE")
    weight_col = col_index(headers, "体重", "WEIGHT")
    height_col = col_index(headers, "身高", "HEIGHT")
    bmi_col = col_index(headers, "BMI")
    perf_col = col_index(headers, "是否检查", "是否测量")

    records = []
    for row in rows_data[1:]:
        subj = parse_subject_id(row[subj_col]) if subj_col is not None else None
        if not subj:
            continue
        if perf_col is not None and norm(row[perf_col]) in ("否", "No", "no"):
            continue
        date = to_date(row[date_col]) if date_col is not None else None
        visit = norm(row[visit_col]) if visit_col is not None else date or ""
        visit_point = extract_visit_point(visit)
        center = norm(row[center_col]) if center_col is not None else (
            norm(row[center_id_col]) if center_id_col is not None else ""
        )

        if weight_col is not None:
            weight = to_float(row[weight_col])
            if weight is not None:
                records.append({
                    "subj": subj, "center": center, "visit": visit, "visit_point": visit_point, "date": date,
                    "test": "体重", "result": weight, "unit": "kg",
                    "lo": None, "hi": None, "cs": "", "desc": "",
                    "desc_link": "", "note": "",
                    "ae_ref": "", "mh_ref": "", "source_sheet": "HW",
                })
        if bmi_col is not None:
            bmi = to_float(row[bmi_col])
            if bmi is not None:
                records.append({
                    "subj": subj, "center": center, "visit": visit, "visit_point": visit_point, "date": date,
                    "test": "BMI", "result": bmi, "unit": "kg/m²",
                    "lo": None, "hi": None, "cs": "", "desc": "",
                    "desc_link": "", "note": "",
                    "ae_ref": "", "mh_ref": "", "source_sheet": "HW",
                })

    return records


def parse_ae_mh_cm_rows(ws, sheet_type):
    """Parse AE, MH, or CM rows."""
    rows_data = list(ws.iter_rows(values_only=True))
    if not rows_data:
        return []
    headers_raw = [norm(h) for h in rows_data[0]]
    headers = {h: i for i, h in enumerate(headers_raw) if h}

    subj_col = col_index(headers, "受试者编号", "受试者", "SUBJID")

    if sheet_type == "AE":
        term_col = col_index(headers, "不良事件名称", "AE名称", "AETERM")
        start_col = col_index(headers, "开始日期", "最早开始日期", "AESTDAT")
        end_col = col_index(headers, "结束日期", "AEENDAT")
        sev_col = col_index(headers, "严重程度", "AESEV", "最严重程度")
        rel_col = col_index(headers, "与试验药物的关系", "AEREL")
        out_col = col_index(headers, "转归", "AEOUT")
        ser_col = col_index(headers, "是否为严重不良事件", "AESER")
    elif sheet_type == "MH":
        term_col = col_index(headers, "疾病名称", "MHTERM")
        start_col = col_index(headers, "开始日期", "MHSTDAT")
        end_col = col_index(headers, "结束日期", "MHENDAT")
        ongoing_col = col_index(headers, "是否持续", "MHONGO")
    elif sheet_type == "CM":
        term_col = col_index(headers, "药物名称", "CMTRT")
        start_col = col_index(headers, "开始日期", "CMSTDAT")
        end_col = col_index(headers, "结束日期", "CMENDAT")
        reason_col = col_index(headers, "用药原因", "CMINDC")
        route_col = col_index(headers, "给药途径", "CMROUTE")

    records = []
    for row in rows_data[1:]:
        subj = parse_subject_id(row[subj_col]) if subj_col is not None else None
        if not subj:
            continue

        term = norm(row[term_col]) if term_col is not None else ""
        if not term:
            continue

        rec = {
            "subj": subj,
            "term": term,
        }
        if sheet_type == "AE":
            rec.update({
                "start": to_date(row[start_col]) if start_col is not None else None,
                "end": to_date(row[end_col]) if end_col is not None else None,
                "severity": norm(row[sev_col]) if sev_col is not None else "",
                "relationship": norm(row[rel_col]) if rel_col is not None else "",
                "outcome": norm(row[out_col]) if out_col is not None else "",
                "serious": norm(row[ser_col]) if ser_col is not None else "",
            })
        elif sheet_type == "MH":
            rec.update({
                "start": to_date(row[start_col]) if start_col is not None else None,
                "end": to_date(row[end_col]) if end_col is not None else None,
                "ongoing": norm(row[ongoing_col]) if ongoing_col is not None else "",
            })
        elif sheet_type == "CM":
            rec.update({
                "start": to_date(row[start_col]) if start_col is not None else None,
                "end": to_date(row[end_col]) if end_col is not None else None,
                "reason": norm(row[reason_col]) if reason_col is not None else "",
                "route": norm(row[route_col]) if route_col is not None else "",
            })
        records.append(rec)
    return records


def parse_sv_rows(ws):
    """Parse subject visit records to map subject -> visits."""
    rows_data = list(ws.iter_rows(values_only=True))
    if not rows_data:
        return {}
    headers_raw = [norm(h) for h in rows_data[0]]
    headers = {h: i for i, h in enumerate(headers_raw) if h}

    subj_col = col_index(headers, "受试者编号", "受试者", "SUBJID")
    visit_col = col_index(headers, "数据节")
    if visit_col is None:
        visit_col = match_col(headers, r"访视|VISIT")
    date_col = col_index(headers, "访视日期", "VISDAT")

    visits = {}
    for row in rows_data[1:]:
        subj = parse_subject_id(row[subj_col]) if subj_col is not None else None
        if not subj:
            continue
        visit = norm(row[visit_col]) if visit_col is not None else ""
        date = to_date(row[date_col]) if date_col is not None else None
        if subj not in visits:
            visits[subj] = []
        visits[subj].append({"visit": visit, "date": date})
    return visits


def parse_group_table(filepath):
    """Parse subject grouping table."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    groups = {}
    for name in wb.sheetnames:
        if "summary" in name.lower() or "toc" in name.lower():
            continue
        ws = wb[name]
        rows_data = list(ws.iter_rows(values_only=True))
        if not rows_data:
            continue
        headers_raw = [norm(h) for h in rows_data[0]]
        headers = {h: i for i, h in enumerate(headers_raw) if h}

        subj_col = col_index(headers, "受试者编号", "受试者", "SUBJID")
        group_col = col_index(headers, "治疗组", "分组", "Group", "ARM")
        if subj_col is None or group_col is None:
            continue

        for row in rows_data[1:]:
            subj = parse_subject_id(row[subj_col])
            group = norm(row[group_col])
            if subj and group:
                groups[subj] = group
    wb.close()
    return groups


def classify_visit(visit_str):
    """Classify visit as screening, baseline, or post-baseline."""
    v = norm(visit_str).upper()

    # Screening
    if any(k in v for k in ["筛选", "SCR", "SCREENING", "D-", "导入期", "磨合期", "RUN-IN"]):
        return "screening"

    # Baseline / D1
    if re.search(r'(^|[（(])D1\b|[（(]基线[）)]|^基线$|BASELINE', v):
        return "baseline"

    # Post-baseline (has D+number)
    if re.search(r'D\d+', v):
        return "post"

    # Treatment end / EOS
    if any(k in v for k in ["治疗结束", "EOS", "EOT", "END OF", "结束"]):
        return "post"

    return "post"


def get_visit_rank(visit_str):
    """Get numeric rank for visit ordering."""
    v = norm(visit_str)
    if classify_visit(v) == "screening":
        return 0
    if classify_visit(v) == "baseline":
        return 1
    m = re.search(r'D(\d+)', v)
    if m:
        return int(m.group(1))
    # EOS = very high
    return 999


def parse_listing(xlsx_path):
    """Main function: Parse an entire Data Listing Excel file."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    result = {
        "lab_records": [],      # All lab/vital/ECG/weight records
        "ae_records": [],       # AE records
        "mh_records": [],       # MH records
        "cm_records": [],       # CM records
        "sv_records": {},       # Subject visits
        "subjects": set(),      # All subject IDs
        "centers": set(),       # All center names
        "test_names": set(),    # All unique test names
        "test_sources": {},     # Which source each test comes from
        "sheet_summary": {},    # Summary of each sheet parsed
        "warnings": [],         # Warnings / missing data
    }

    for name in wb.sheetnames:
        ws = wb[name]
        rows_data = list(ws.iter_rows(values_only=True))
        if not rows_data:
            continue
        headers_list = [h for h in rows_data[0] if h is not None]
        sheet_type = classify_sheet(name, headers_list)

        if sheet_type is None:
            continue

        result["sheet_summary"][name] = {
            "type": sheet_type,
            "rows": len(rows_data) - 1,
        }

        if sheet_type in ("LB_HEM", "LB_CHEM", "LB_URI", "LB_OTHER"):
            records = parse_lab_rows(ws, sheet_type)
            result["lab_records"].extend(records)
        elif sheet_type == "VS":
            records = parse_vs_rows(ws)
            result["lab_records"].extend(records)
        elif sheet_type == "EG":
            records = parse_eg_rows(ws)
            result["lab_records"].extend(records)
        elif sheet_type == "HW":
            records = parse_hw_rows(ws)
            result["lab_records"].extend(records)
        elif sheet_type == "AE":
            result["ae_records"] = parse_ae_mh_cm_rows(ws, "AE")
        elif sheet_type == "MH":
            result["mh_records"] = parse_ae_mh_cm_rows(ws, "MH")
        elif sheet_type == "CM":
            result["cm_records"] = parse_ae_mh_cm_rows(ws, "CM")
        elif sheet_type == "SV":
            result["sv_records"] = parse_sv_rows(ws)

    wb.close()

    # Deduplicate and collect metadata
    for rec in result["lab_records"]:
        result["subjects"].add(rec["subj"])
        if rec.get("center"):
            result["centers"].add(rec["center"])
        result["test_names"].add(rec["test"])
        if rec["test"] not in result["test_sources"]:
            result["test_sources"][rec["test"]] = rec.get("source_sheet", "unknown")

    # Check for missing data
    has_lab_hem = any(s["type"] == "LB_HEM" for s in result["sheet_summary"].values())
    has_lab_chem = any(s["type"] == "LB_CHEM" for s in result["sheet_summary"].values())
    has_vs = any(s["type"] == "VS" for s in result["sheet_summary"].values())
    has_eg = any(s["type"] == "EG" for s in result["sheet_summary"].values())
    has_hw = any(s["type"] == "HW" for s in result["sheet_summary"].values())

    if not has_vs:
        result["warnings"].append("未发现生命体征(VS)相关sheet。将不分析生命体征指标。")
    if not has_eg:
        result["warnings"].append("未发现心电图(EG)相关sheet。将不分析心电图指标。")
    if not has_hw:
        result["warnings"].append("未发现身高体重(HW)相关sheet。将不分析体重指标。")
    if not has_lab_hem and not has_lab_chem:
        result["warnings"].append("未发现实验室检查(LB_HEM/LB_CHEM)相关sheet。")

    # Filter out tests only present in screening/baseline
    test_visit_phases = defaultdict(set)
    for rec in result["lab_records"]:
        phase = classify_visit(rec.get("visit", ""))
        test_visit_phases[rec["test"]].add(phase)

    excluded_tests = []
    for test_name, phases in test_visit_phases.items():
        post_phases = phases - {"screening", "baseline"}
        if not post_phases:
            excluded_tests.append(test_name)
            result["warnings"].append(
                f"指标「{test_name}」仅在筛选期/基线期有数据，基线后无数据，已排除分析。"
            )

    if excluded_tests:
        result["lab_records"] = [
            rec for rec in result["lab_records"]
            if rec["test"] not in excluded_tests
        ]
        result["test_names"] = {t for t in result["test_names"] if t not in excluded_tests}

    # Filter out single-occurrence tests (only 1 record across ALL subjects)
    test_counts = Counter(rec["test"] for rec in result["lab_records"])
    single_tests = [t for t, c in test_counts.items() if c <= 1]
    if single_tests:
        result["warnings"].append(
            f"以下指标在整个研究中仅有一次检测记录，已排除分析: {', '.join(single_tests)}"
        )
        result["lab_records"] = [
            rec for rec in result["lab_records"] if rec["test"] not in single_tests
        ]
        result["test_names"] = {t for t in result["test_names"] if t not in single_tests}

    # Build visit label mapping from SV records
    # Maps (subj, date) → visit name for enriching records with missing visit names
    visit_label_map = {}
    for subj, visits in result.get("sv_records", {}).items():
        for v in visits:
            d = v.get("date")
            vn = v.get("visit", "")
            if d and vn:
                key = (subj, d)
                if key not in visit_label_map:
                    visit_label_map[key] = vn

    # Enrich visit labels: if visit is just a date, try to map to actual visit name
    enriched_count = 0
    for rec in result["lab_records"]:
        visit = rec.get("visit", "")
        date = rec.get("date", "")
        sv_key = (rec["subj"], date) if date else None
        # If visit looks like a date-only or is empty, use SV mapping
        if sv_key and sv_key in visit_label_map:
            sv_visit = visit_label_map[sv_key]
            if not visit or re.match(r'\d{4}-\d{2}-\d{2}', visit):
                rec["visit"] = sv_visit
                enriched_count += 1

    if enriched_count > 0:
        result["warnings"].append(
            f"通过SV访视表补充了 {enriched_count} 条记录的访视名称。"
        )

    # Detect CS="异常有临床意义" but no explanation
    cs_missing_desc = []
    for rec in result["lab_records"]:
        cs = norm(rec.get("cs", ""))
        desc = norm(rec.get("desc", ""))
        desc_link = norm(rec.get("desc_link", ""))
        note = norm(rec.get("note", ""))
        combined = desc or desc_link or note
        if any(k in cs for k in ("有临床意义", "CS")) and "无临床意义" not in cs and not combined:
            cs_missing_desc.append(
                {"subj": rec["subj"], "test": rec["test"], "visit": rec.get("visit", ""), "sheet": rec.get("source_sheet", "")}
            )
    if cs_missing_desc:
        samples = cs_missing_desc[:5]
        sample_strs = [f"{s['subj']}/{s['test']}({s['visit']})" for s in samples]
        result["cs_missing_desc"] = cs_missing_desc
        result["warnings"].append(
            f"发现 {len(cs_missing_desc)} 条评估为CS（异常有临床意义）但临床意义解释为空的记录。"
            f"示例: {', '.join(sample_strs)}。"
            "请确认data listing中哪一列可获知临床意义解释（如'动态链接'列中的'AE：xxx'/'MH：xxx'等）。"
        )

    # Sort records by subject, visit rank, date
    result["lab_records"].sort(key=lambda r: (
        r["subj"],
        get_visit_rank(r.get("visit", "")),
        r.get("date") or "",
    ))

    # Convert sets to lists for JSON
    result["subjects"] = sorted(result["subjects"])
    result["centers"] = sorted(result["centers"])
    result["test_names"] = sorted(result["test_names"])

    # Summary stats
    result["stats"] = {
        "total_subjects": len(result["subjects"]),
        "total_centers": len(result["centers"]),
        "total_test_names": len(result["test_names"]),
        "total_lab_records": len(result["lab_records"]),
        "total_ae_records": len(result["ae_records"]),
        "total_mh_records": len(result["mh_records"]),
        "total_cm_records": len(result["cm_records"]),
    }

    return result


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python parse_listing.py <xlsx_file> [output.json]")
        sys.exit(1)

    data = parse_listing(sys.argv[1])
    out_path = sys.argv[2] if len(sys.argv) > 2 else "parsed_data.json"
    Path(out_path).write_text(
        json.dumps(data, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8"
    )
    print(f"Parsed: {data['stats']}")
    print(f"Warnings: {len(data['warnings'])}")
    for w in data['warnings']:
        print(f"  ⚠ {w}")
    print(f"Output: {out_path}")
